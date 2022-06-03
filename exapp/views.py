import re

import xlrd as xlrd
import xlwt

from django.http import HttpResponse
from django.shortcuts import render


# Create your views here.
from openpyxl import load_workbook

from dbmanageapp.models import UploadDb, UploadDbName, MarketingList, DbMemo
from datetime import datetime

def exdown(request):
    response = HttpResponse(content_type="application/vnd.ms-excel")
    response["Content-Disposition"] = 'attachment;filename*=UTF-8\'\'example.xls'
    wb = xlwt.Workbook(encoding='ansi')  # encoding은 ansi로 해준다.
    ws = wb.add_sheet('sheet1')  # 시트 추가

    row_num = 0
    col_names = ['db_id','db_name','db_mkname','db_phone','db_member','db_age','db_sex','db_inv','db_manager','db_manager_nick','db_status','db_paidprice','db_paidstatus','db_lastpaiddate','db_date']
    for idx, col_name in enumerate(col_names):
        ws.write(row_num, idx, col_name)

    rows = UploadDb.objects.all().values_list()
    print(rows)
    for row in rows:
        row = list(row)
        row[13] = row[13].strftime("%Y/%m/%d/ %H:%M-%S.%f")
        row[14] = row[14].strftime("%Y/%m/%d/ %H:%M-%S.%f")

        row_num += 1
        for col_num, attr in enumerate(row):
            ws.write(row_num, col_num, attr)
    wb.save(response)

    return response

def ex_setting(request):
    response = HttpResponse(content_type="application/vnd.ms-excel")
    response["Content-Disposition"] = 'attachment;filename*=UTF-8\'\'example.xls'
    wb = xlwt.Workbook(encoding='ansi')  # encoding은 ansi로 해준다.
    ws = wb.add_sheet('sheet1')  # 시트 추가

    row_num = 0
    col_names = ['db_id','db_name','db_mkname','db_phone','db_member','db_age','db_sex','db_inv','db_manager','db_manager_nick','db_status','db_paidprice','db_paidstatus','db_lastpaiddate','db_date']
    for idx, col_name in enumerate(col_names):
        ws.write(row_num, idx, col_name)

    rows = UploadDb.objects.all().values_list()
    for row in rows:

        # print(row)

        row = list(row)
        row[13] = row[13].strftime("%Y/%m/%d/ %H:%M-%S.%f")
        row[14] = row[14].strftime("%Y/%m/%d/ %H:%M-%S.%f")
        # row[14] = 000000
        row_num += 1

    testdate = "2022/05/22/ 16:29-47.920026"

    datetime_string = "2022/05/22/ 16:29-47.920026"
    datetime_format = "%Y/%m/%d/ %H:%M-%S.%f"
    datetime_result = datetime.strptime(datetime_string, datetime_format)
    # print(datetime_result)

    if request.method == 'POST':
        dbname_id = request.POST.get('dbname')
        print(dbname_id)
        chk_dbname = UploadDbName.objects.get(id=dbname_id)
        print(chk_dbname)
        print(chk_dbname.dbn_mkname)


        files = request.FILES['uploadex']
        load_wb = load_workbook(files, data_only=True)
        load_ws = load_wb['Sheet1']
        dblist = []
        for row in load_ws.rows:
            row_value = []

            rcount = 0
            for cell in row:
                rcount += 1
                cellval = cell.value

                if rcount == 4 or rcount == 5:
                    print(type(cellval))
                    cellval = str(cellval)
                    cellval = cellval.zfill(11)

                if cellval == 'None' or not cellval:
                    cellval = ""
                row_value.append(cellval)


            dblist.append(row_value)

        print(dblist)
        for dd in dblist:
            print(dd[11])


            datetime_string = dd[13]
            datetime_format = "%Y/%m/%d/ %H:%M-%S.%f"
            dd[13] = datetime.strptime(datetime_string, datetime_format)

            datetime_string = dd[14]
            datetime_format = "%Y/%m/%d/ %H:%M-%S.%f"
            dd[14] = datetime.strptime(datetime_string, datetime_format)

            if dd[11] == 0 or dd[11] == '' or dd[11] is None:
                set_paidprice = 0



            UploadDb.objects.create(db_name=chk_dbname, db_mkname=chk_dbname.dbn_mkname, db_phone=dd[3], db_member=dd[4], db_age=dd[5], db_sex=dd[6],db_inv=dd[7],db_manager=dd[8],db_manager_nick=dd[9],db_status=dd[10],db_paidprice=set_paidprice,db_paidstatus=dd[12],db_lastpaiddate=dd[13],db_date=dd[14])


    return render(request, 'exapp/ex_seton.html')


def memo_set(request):
    chk_db_list = DbMemo.objects.all()
    set_count = len(chk_db_list)

    if request.method == "POST":
        for dlist in chk_db_list:
            if dlist.dm_mamager == "":
                dlist.dm_mamager = dlist.dm_chkdb.db_manager
                dlist.save()

    return render(request, 'exapp/ex_seton.html', {'set_count':set_count})
